from pathlib import Path
import csv
from openpyxl import load_workbook

def is_blank(value):
    """
    空欄判定
    None、空文字、スペースのみを空欄とみなす
    """
    return value is None or str(value).strip() == ""

def calc_average(values):
    """
    平均値を計算する関数
    空リストの場合は None を返す
    """
    valid_values = []

    for value in values:
        if value is not None:
            valid_values.append(value)

    if len(valid_values) == 0:
        return None
  
    return sum(valid_values) / len(valid_values)

def to_float(value):
    """
    数値に変換する関数
    空欄または数値変換できない値は None を返す
    """
    if is_blank(value):
        return None

    try:
        return float(value)
    except ValueError:
        return None

def format_decimal(value, digits):
    """
    小数点以下の桁数をそろえる関数
    """
    value_float = to_float(value)

    if value_float is None:
        return ""

    return f"{value_float:.{digits}f}"

def format_survey_datetime(value):
    """
    調査日時を表示用の文字列に整える関数
    Excelの値をそのまま使い、空欄の場合は空文字を返す
    """
    if is_blank(value):
        return ""

    return str(value).strip()

def is_unused_tree_row(tree_species, tree_height, 
    branch_height, dbh, abnormal_type, damage_type):

    """
    毎木調査の未使用行かどうかを判定する関数
    番号は最初から入力されている可能性があるため、判定に使わない
    """
    return (
        is_blank(tree_species)
        and is_blank(tree_height)
        and is_blank(branch_height)
        and is_blank(dbh)
        and is_blank(abnormal_type)
        and is_blank(damage_type)
    )


# 基本情報欄をチェックする
def check_basic_info(ws, sheet_name):
    """
    基本情報欄をチェックする関数
    """

    errors = []

    basic_rules = [
        ("調査ID", "B2"),
        ("地域名", "D2"),
        ("調査日時", "F2"),
        ("天気", "H2"),
        ("記帳者", "B3"),
        ("対象樹種", "D3"),
        ("面積", "F3"),
        ("斜面位置", "H3"),
        ("斜面方位", "B4"),
        ("傾斜度", "D4"),
        ("緯度", "F4"),
        ("経度", "H4"),
    ]
    # 未入力のエラーを抽出
    for item_name, cell_address in basic_rules:
        value = ws[cell_address].value

        if is_blank(value):
            errors.append({
                "sheet_name": sheet_name,
                "row_no": "基本情報",
                "item_name": item_name,
                "error_type": "未入力",
                "error_message": f"{item_name}が未入力です"
            })
    # 数値（傾斜度、経度、緯度）のエラーを抽出
    inclination = ws["D4"].value
    lat = ws["F4"].value
    lon = ws["H4"].value

    if not is_blank(inclination) and to_float(inclination) is None:
        errors.append({
            "sheet_name": sheet_name,
            "row_no": "基本情報",
            "item_name": "傾斜度",
            "error_type": "数値エラー",
            "error_message": "傾斜度が数値ではありません"
        })
    
    if inclination < 0 or inclination > 90:
       errors.append({
            "sheet_name": sheet_name,
            "row_no": "基本情報",
            "item_name": "傾斜度",
            "error_type": "数値エラー",
            "error_message": "傾斜度が0〜90度の範囲外です"
        })
       
    if not is_blank(lat) and to_float(lat) is None:
        errors.append({
            "sheet_name": sheet_name,
            "row_no": "基本情報",
            "item_name": "緯度",
            "error_type": "数値エラー",
            "error_message": "緯度が数値ではありません"
        })

    if not is_blank(lon) and to_float(lon) is None:
        errors.append({
            "sheet_name": sheet_name,
            "row_no": "基本情報",
            "item_name": "経度",
            "error_type": "数値エラー",
            "error_message": "経度が数値ではありません"
        })

    return errors

# 毎木調査欄をチェックする
def check_tree_rows(ws, sheet_name):
    """
    毎木調査欄をチェックする関数
    """

    errors = []

    start_row = 7
    end_row = 76

    columns = {
        "番号": 1,
        "樹種": 2,
        "樹高_m": 3,
        "枝下高_m": 4,
        "胸高直径_cm": 5,
        "異常区分": 6,
        "被害区分": 7,
        "備考": 8,
    }

    required_items = [
        "番号",
        "樹種",
        "樹高_m",
        "枝下高_m",
        "胸高直径_cm",
        "異常区分",
        "被害区分",
    ]

    for row in range(start_row, end_row + 1):

        tree_species = ws.cell(row=row, column=columns["樹種"]).value
        tree_height = ws.cell(row=row, column=columns["樹高_m"]).value
        branch_height = ws.cell(row=row, column=columns["枝下高_m"]).value
        dbh = ws.cell(row=row, column=columns["胸高直径_cm"]).value
        abnormal_type = ws.cell(row=row, column=columns["異常区分"]).value
        damage_type = ws.cell(row=row, column=columns["被害区分"]).value

        # 番号以外がすべて空欄なら未使用行としてスキップする
        if (
            is_blank(tree_species)
            and is_blank(tree_height)
            and is_blank(branch_height)
            and is_blank(dbh)
            and is_blank(abnormal_type)
            and is_blank(damage_type)
        ):
            continue

        # 必須項目チェック
        for item_name in required_items:
            col = columns[item_name]
            value = ws.cell(row=row, column=col).value

            if is_blank(value):
                errors.append({
                    "sheet_name": sheet_name,
                    "row_no": row,
                    "item_name": item_name,
                    "error_type": "未入力",
                    "error_message": f"{item_name}が未入力です"
                })

        # 数値変換
        tree_height_value = to_float(tree_height)
        branch_height_value = to_float(branch_height)
        dbh_value = to_float(dbh)

        # 数値チェック
        if not is_blank(tree_height) and tree_height_value is None:
            errors.append({
                "sheet_name": sheet_name,
                "row_no": row,
                "item_name": "樹高_m",
                "error_type": "数値エラー",
                "error_message": "樹高_mが数値ではありません"
            })

        if not is_blank(branch_height) and branch_height_value is None:
            errors.append({
                "sheet_name": sheet_name,
                "row_no": row,
                "item_name": "枝下高_m",
                "error_type": "数値エラー",
                "error_message": "枝下高_mが数値ではありません"
            })

        if not is_blank(dbh) and dbh_value is None:
            errors.append({
                "sheet_name": sheet_name,
                "row_no": row,
                "item_name": "胸高直径_cm",
                "error_type": "数値エラー",
                "error_message": "胸高直径_cmが数値ではありません"
            })

        # 胸高直径が0かチェック
        if dbh_value is not None and dbh_value == 0:
            errors.append({
                "sheet_name": sheet_name,
                "row_no": row,
                "item_name": "胸高直径_cm",
                "error_type": "数値異常",
                "error_message": "胸高直径_cmが0です"
            })

        # 枝下高が樹高より大きくないかチェック
        if (
            tree_height_value is not None
            and branch_height_value is not None
            and branch_height_value > tree_height_value
        ):
            errors.append({
                "sheet_name": sheet_name,
                "row_no": row,
                "item_name": "枝下高_m",
                "error_type": "数値異常",
                "error_message": "枝下高_mが樹高_mより大きいです"
            })

    return errors


def create_qgis_summary_row(ws, sheet_name):
    """
    QGIS結合用CSVに出力する1地点分の集計行を作成する
    """

    # 基本情報
    plot_id = ws["B2"].value # 調査ID
    area_name = ws["D2"].value # 地域名
    survey_date = ws["F2"].value # 調査日時
    weather = ws["H2"].value # 天気
    writer = ws["B3"].value # 記帳者
    tree_species = ws["D3"].value # 対象樹種
    area = ws["F3"].value # 面積_㎡
    slope_position = ws["H3"].value # 斜面位置
    slope_aspect = ws["B4"].value # 斜面方位
    inclination = ws["D4"].value # 傾斜度
    latitude = ws["F4"].value # 緯度
    longitude = ws["H4"].value # 経度
    
    # 毎木調査の開始行
    start_row = 7

    columns = {
        "番号": 1,
        "樹種": 2,
        "樹高_m": 3,
        "枝下高_m": 4,
        "胸高直径_cm": 5,
        "異常区分": 6,
        "被害区分": 7,
        "備考": 8,
    }

    tree_heights = []
    branch_heights = []
    dbhs = []
    tree_count = 0

    for row in range(start_row, ws.max_row + 1):
        tree_species_value = ws.cell(row=row, column=columns["樹種"]).value

        # 樹種が空欄なら、その行は立木データとして扱わない
        if is_blank(tree_species_value):
            continue

        tree_count += 1

        tree_height = to_float(ws.cell(row=row, column=columns["樹高_m"]).value)
        branch_height = to_float(ws.cell(row=row, column=columns["枝下高_m"]).value)
        dbh = to_float(ws.cell(row=row, column=columns["胸高直径_cm"]).value)

        if tree_height is not None:
            tree_heights.append(tree_height)

        if branch_height is not None:
            branch_heights.append(branch_height)

        if dbh is not None:
            dbhs.append(dbh)

    summary_row = {
        "調査ID": plot_id,
        "シート名": sheet_name,
        "地域名": area_name,
        "調査日時": format_survey_datetime(survey_date),
        "天気": weather,
        "記帳者": writer,
        "対象樹種": tree_species,
        "面積_㎡": area,
        "斜面位置": slope_position,
        "斜面方位": slope_aspect,
        "傾斜度": format_decimal(inclination,1),
        "緯度": format_decimal(latitude,6),
        "経度": format_decimal(longitude,6),
        
        "平均樹高_m": format_decimal(calc_average(tree_heights),1),
        "平均枝下高_m": format_decimal(calc_average(branch_heights),1),
        "平均胸高直径_cm": format_decimal(calc_average(dbhs),1),
        "立木本数": tree_count,
    }

    return summary_row


def write_qgis_csv(qgis_rows, output_path):
    """
    QGIS結合用CSVを出力する関数
    """

    fieldnames = [
        "調査ID",
        "シート名",
        "地域名",
        "調査日時",
        "天気",
        "記帳者",
        "対象樹種",
        "面積_㎡",
        "斜面位置",
        "斜面方位",
        "傾斜度",
        "緯度",
        "経度",
        "平均樹高_m",
        "平均枝下高_m",
        "平均胸高直径_cm",
        "立木本数",
    ]

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(qgis_rows)

def main():
    base_dir = Path(r"E:\protopype\01_survey_tool")

    input_file = base_dir / "01_input" / "sample_forest_survey_3plots.xlsx"
    output_file = base_dir / "03_output" / "error_log.csv"
    qgis_csv_path = base_dir / "03_output" / "qgis_plot_summary.csv"

    wb = load_workbook(input_file, data_only=True)

    target_sheets = ["P001", "P002", "P003"]

    all_errors = []

    for sheet_name in target_sheets:
        ws = wb[sheet_name]

        basic_errors = check_basic_info(ws, sheet_name)
        tree_errors = check_tree_rows(ws, sheet_name)

        all_errors.extend(basic_errors)
        all_errors.extend(tree_errors)

    output_file.parent.mkdir(parents=True, exist_ok=True)

    with open(output_file, "w", newline="", encoding="utf-8-sig") as f:
        fieldnames = [
            "sheet_name",
            "row_no",
            "item_name",
            "error_type",
            "error_message",
        ]

        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_errors)

        print(f"チェック完了: {output_file}")
        print(f"エラー件数: {len(all_errors)}")

         # エラーがある場合は、QGIS用CSVを出力しない
        if all_errors:
            print("エラーがあるため、qgis_plot_summary.csv は出力しません。")
            print("error_log.csv を確認して、Excelを修正してから再実行してください。")
            return

        # エラーがない場合だけ、QGIS用CSVを作成する
        qgis_rows = []

        for ws in wb.worksheets:
            sheet_name = ws.title
            summary_row = create_qgis_summary_row(ws, sheet_name)
            qgis_rows.append(summary_row)

            write_qgis_csv(qgis_rows, qgis_csv_path)

if __name__ == "__main__":
    main()