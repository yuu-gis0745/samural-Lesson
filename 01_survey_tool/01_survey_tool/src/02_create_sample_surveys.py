from pathlib import Path
from openpyxl import load_workbook
import random

BASE_DIR = Path(__file__).resolve().parent.parent
template_file = BASE_DIR / "02_master" / "template_forest_survey.xlsx"
output_file = BASE_DIR / "01_input" / "sample_forest_survey_3plots.xlsx"

# =========================
# 設定
# =========================
TREE_COUNT = 30  # 1地点あたりの入力本数

TREE_SPECIES = ["スギ", "ヒノキ", "アカマツ", "カラマツ"]
WEATHER_LIST = ["晴れ", "曇り", "雨","雪"]
ASPECT_LIST = ["北", "北東", "東", "南東", "南", "南西", "西", "北西", "平坦"]
SLOPE_POSITION_LIST = ["山頂部", "上部斜面", "中部斜面", "下部斜面", "谷部", "平坦地"]
ABNORMAL_CATEGOLY_LIST = ["なし", "枯損木", "先折れ", "二股木", "傾斜木", "空洞", "腐朽", "変色","その他"]
DAMAGE_CATEGOLY_LIST = ["なし", "獣害", "病虫害", "風雪害", "その他"]

# 毎木調査欄の開始行
START_ROW = 7

# 毎木調査欄の列位置
COL_NO = "A"
COL_SPECIES = "B"
COL_HEIGHT = "C"
COL_BRANCH_HEIGHT = "D"
COL_DBH = "E"
COL_ABNORMAL = "F"
COL_DAMAGE = "G"
COL_NOTE = "H"


# =========================
# 共通処理
# =========================

def set_basic_info(ws, plot_id: str):
    """基本情報を入力する"""

    basic_data = {
        "P001": {
            "調査ID": "P001",
            "地域名": "A地区",
            "調査日時": '2026/5/17 12:30～13:30',
            "天気": "晴れ",
            "記帳者": "侍",
            "対象樹種": "スギ",
            "面積": 400,
            "斜面方位": "北東",
            "斜面位置": "中部斜面",
            "傾斜度": 25,
            "緯度": 35.681236,
            "経度": 139.767125,
        },
        "P002": {
            "調査ID": "P002",
            "地域名": "B地区",
            "調査日時": '2026/5/17 14:10～15:00',
            "天気": "曇り",
            "記帳者": "侍",
            "対象樹種": "スギ",
            "面積": 400,
            "斜面方位": "南西",
            "斜面位置": "上部斜面",
            "傾斜度": 30,
            "緯度": 35.682100,
            "経度": 139.768200,
        },
        "P003": {
            "調査ID": "P003",
            "地域名": "C地区",
            "調査日時": '2026/5/18 9:30～10:30',
            "天気": "晴れ",
            "記帳者": "侍",
            "対象樹種": "スギ",
            "面積": 200,
            "斜面方位": "東",
            "斜面位置": "下部斜面",
            "傾斜度": 18,
            "緯度": 35.683000,
            "経度": 139.769000,
        },
    }

    d = basic_data[plot_id]

    # ※セル位置は、現在の森林調査表に合わせて必要に応じて調整してください
    ws["B2"] = d["調査ID"]
    ws["D2"] = d["地域名"]
    ws["F2"] = d["調査日時"]
    ws["H2"] = d["天気"]

    ws["B3"] = d["記帳者"]
    ws["D3"] = d["対象樹種"]
    ws["F3"] = d["面積"]
    ws["H3"] = d["斜面位置"]
    ws["B4"] = d["斜面方位"]
    ws["D4"] = d["傾斜度"]
    ws["F4"] = d["緯度"]
    ws["H4"] = d["経度"]


def create_tree_data(ws, tree_count: int):
    """毎木調査データをランダム作成する"""

    for i in range(tree_count):
        row = START_ROW + i

        tree_no = i + 1
        species = "スギ"

        height = round(random.uniform(8.0, 30.0), 1)
        branch_height = round(random.uniform(1.0, height * 0.7), 1)
        dbh = round(random.uniform(8.0, 60.0), 1)

        ws[f"{COL_NO}{row}"] = tree_no
        ws[f"{COL_SPECIES}{row}"] = species
        ws[f"{COL_HEIGHT}{row}"] = height
        ws[f"{COL_BRANCH_HEIGHT}{row}"] = branch_height
        ws[f"{COL_DBH}{row}"] = dbh
        ws[f"{COL_ABNORMAL}{row}"] = "なし"
        ws[f"{COL_DAMAGE}{row}"] = "なし"
        ws[f"{COL_NOTE}{row}"] = ""


def add_errors(ws, plot_id: str):
    """P002・P003に意図的なエラーを入れる"""

    if plot_id == "P002":
        # 基本情報エラー
        ws["H2"] = ""   # 天気空白
        ws["B4"] = ""   # 斜面方位空白

        # 毎木調査エラー
        ws["E10"] = 0       # 胸高直径0
        ws["D12"] = 25.0    # 枝下高
        ws["C12"] = 15.0    # 樹高 → 枝下高が樹高より大きい

    elif plot_id == "P003":
        # 基本情報エラー
        ws["F4"] = ""   # 緯度空白

        # 毎木調査エラー
        ws["C9"] = ""       # 樹高空白
        ws["D9"] = 5.0      # 枝下高あり

        ws["F15"] = ""      # 異常区分空白
        ws["G16"] = ""      # 被害区分空白


def apply_number_format(ws):
    """表示形式を設定する"""

    # 基本情報
    ws["F2"].number_format = "yyyy/m/d h:mm"      # 調査日
    ws["F3"].number_format = "000"      # 面積
    ws["D4"].number_format = "0"         # 傾斜度
    ws["F4"].number_format = "0.000000"  # 緯度
    ws["H4"].number_format = "0.000000"  # 経度

    # 毎木調査欄
    for row in range(START_ROW, START_ROW + TREE_COUNT):
        ws[f"{COL_HEIGHT}{row}"].number_format = "0.0"
        ws[f"{COL_BRANCH_HEIGHT}{row}"].number_format = "0.0"
        ws[f"{COL_DBH}{row}"].number_format = "0.0"


# =========================
# メイン処理
# =========================

def main():
    if not template_file.exists():
        raise FileNotFoundError(f"テンプレートが見つかりません: {template_file}")
    
    # 出力先フォルダがなければ作成
    output_file.parent.mkdir(parents=True, exist_ok=True)

    plot_ids = ["P001", "P002", "P003"]

    wb = load_workbook(template_file)

    # テンプレートの最初のシートを取得
    template_ws = wb.active

    for index, plot_id in enumerate(plot_ids):
        if index == 0:
            ws = template_ws
            ws.title = plot_id
        else:
            ws = wb.copy_worksheet(template_ws)
            ws.title = plot_id

        set_basic_info(ws, plot_id)
        create_tree_data(ws, TREE_COUNT)
        add_errors(ws, plot_id)
        apply_number_format(ws)

    wb.save(output_file)
    print(f"作成完了: {output_file}")

if __name__ == "__main__":
    main()