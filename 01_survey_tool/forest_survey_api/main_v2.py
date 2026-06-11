from fastapi import FastAPI, UploadFile, File, Request
from fastapi.responses import Response
from openpyxl import load_workbook
from pathlib import Path
import io
import csv
import json
import os
from google.cloud import vision
import fitz  # pymupdf（PDFをページごとに画像化するために使う）

# ==============================================================================================
# 入力されたxlsxから、以下の処理を実行するコードです。
# ・エラー判定のルール設定
# ・OCRによるpdf読み取り
# ・エラー判定
# ・既存のエラー判別コードにないエラー候補を追加
# ・エラーログ出力
# ・GIS用のcsv出力
# ==============================================================================================

app = FastAPI()

# masterフォルダのパスを定義する
# 環境変数 MASTER_DIR があればそれを使い、なければ従来の ./master を使う
DEFAULT_MASTER_DIR = Path(__file__).parent / "master"
MASTER_DIR = Path(os.environ.get("MASTER_DIR", DEFAULT_MASTER_DIR))

# Cloud Vision API 認証ファイルのパスを設定する
# 環境変数 GOOGLE_APPLICATION_CREDENTIALS が未設定のときだけ、ローカル既定値を補う
DEFAULT_GCP_KEY = "C:/Users/Owner/Documents/gcp_keys/vision-ocr.json"
os.environ.setdefault("GOOGLE_APPLICATION_CREDENTIALS", DEFAULT_GCP_KEY)


# ==============================================================================================
# エラー判定のルール設定
# ==============================================================================================

def is_blank(value: str) -> bool:
    """
    空欄判定
    None、空文字、スペースのみを空欄とみなす
    """
    # Noneは、str(None)="None"になるため先に判定を行う
    if value is None:
        return True
    
    # 型にこだわらず判定を行うため、全て文字列に一度変換する
    # 半角・全角スペース(u3000)の両方を除去して判定する
    cleaned = str(value).strip().replace("\u3000", "") 
    return cleaned == ""  # ← ここで比較する

def to_float(value:float)-> float:
    """
    数値に変換する関数
    空欄または数値変換できない値は None を返す
    """

    # 空欄であれば、Noneを返す
    if is_blank(value):
        return None
    
    # 例外処理(数値ではないもの(例：abc)があった場合には、Noneを返す)
    # 数値変換できない値(文字列)であってもErrorとせずNoneを戻り値とする
    try: 
        return float(value) # 例外が発生する可能性のある処理
    except ValueError: # 例外型
        return None # 例外が発生したときに実行する処理

def format_decimal(value, digits:float)-> float:
    """
    小数点以下の桁数をそろえる関数
    """

    value_float = to_float(value)
    
    # value_floatがNoneの場合は、空文字を返す
    if value_float is None:
        return ""
    
    # digits：浮動小数点数（float）を指定した桁数でフォーマット
    # フォーマット文字列を戻り値とする（例：digits=1 なら 3.14159 → "3.1"）
    return f"{value_float:.{digits}f}"

def format_survey_datetime(value:str)-> str:
    """
    調査日時を表示用の文字列に整える関数
    Excelの値をそのまま使い、空欄の場合は空文字を返す
    """

    # 調査日時のNone、スペース、空欄判定
    if is_blank(value):
        return ""
    
    # 空欄や空文字、None、スペースがあれば削除する 
    return str(value).strip()

def is_unused_tree_row(tree_species:str, tree_height:float, 
    branch_height:float, dbh:float, abnormal_type:str, damage_type:str)-> str:
    """
    毎木調査の未使用行かどうかを判定する関数
    番号は最初から入力されている可能性があるため、判定に使わない
    """

    return (
        is_blank(tree_species) # 樹種に空欄や空文字、None、スペースがあるか判定
        and is_blank(tree_height) # 樹高(m)に空欄や空文字、None、スペースがあるか判定
        and is_blank(branch_height) # 枝下高(m)に空欄や空文字、None、スペースがあるか判定
        and is_blank(dbh) # 胸高直径(cm)に空欄や空文字、None、スペースがあるか判定
        and is_blank(abnormal_type) # 異常区分に空欄や空文字、None、スペースがあるか判定
        and is_blank(damage_type)) # 被害区分に空欄や空文字、None、スペースがあるか判定

def load_basic_inf_cell_mapping(cell_mapping_file:str)-> str:
    """
    forest_survey_cell_mapping.xlsx から
    基本情報の項目名とセル位置を取得する関数
    """

    wb = load_workbook(cell_mapping_file, data_only=True)
    sheet = wb["基本情報セル対応表"]

    basic_information = []

    # 列名を取り出す
    for row in sheet.iter_rows(
        min_row=3,
        max_row=14,
        min_col=2,
        max_col=4,
        values_only=True): # セル値のみを返すかどうか判定
        item_name = row[0]      # B列：項目名を取得
        cell_address = row[2]   # D列：セル位置を取得
        
        # 項目名もしくはセル位置の値が空白になったらfor文から抜け出す
        if is_blank(item_name) or is_blank(cell_address):
            continue

        basic_information.append((item_name, cell_address))

    return basic_information

def load_tree_inf_cell_mapping(cell_mapping_file:str)-> str:
    """
    forest_survey_cell_mapping.xlsx から
    毎木情報の項目名とセル位置を取得する関数
    """

    wb = load_workbook(cell_mapping_file, data_only=True)
    sheet = wb["毎木調査欄セル対応表"]
    
    tree_information = []
    i = 1

    for row in sheet.iter_rows(
        min_row=2,
        max_row=9,
        min_col=2,
        max_col=4,
        values_only=True): # セル値のみを返すかどうか判定
        columns = row[0] # B列：項目名
        columns_num = i
        i = i + 1
        
        # 項目名もしくはセル位置の値が空白になったらfor文から抜け出す
        if is_blank(columns) or is_blank(columns_num):
            continue

        tree_information.append((columns, columns_num))

    return tree_information

def load_tree_data_cell_mapping(cell_mapping_file:str)-> tuple:
    """
    forest_survey_cell_mapping.xlsx から
    毎木情報の調査結果の開始行と終了行を取得する関数
    """

    wb = load_workbook(cell_mapping_file, data_only=True)
    sheet = wb["毎木調査欄セル対応表"]
    
    for row in sheet.iter_rows(
        min_row=2,
        max_row=9,
        min_col=2,
        max_col=4,
        values_only=True): # セル値のみを返すかどうか判定

        cell_address = row[2]   # D列：セル範囲 例 B7:B76
        
        # None、空文字、空欄があればcontinueする
        if is_blank(cell_address):
            continue

        start_cell = cell_address.split(":")[0]  # 例 B7
        end_cell = cell_address.split(":")[1]    # 例 B76

        start_row = int(start_cell[1:])  # 7
        end_row = int(end_cell[1:])  # 76

        return start_row, end_row

    return None, None

# check_rules_forest_survey.csvに基づいてエラーを表示する
def load_check_rules(check_rules_file: str) -> list:
    """
    check_rules_forest_survey.csv を読み込む関数
    """

    rules = []
    
    # check_rules_fileをutf-8-sig形式で開く。
    with open(check_rules_file, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)

        for row in reader:
            rules.append({
                "rule_id": row["rule_id"],
                "category": row["category"],
                "check_item": row["check_item"],
                "target_column": row["target_column"],
                "condition": row["condition"],
                "severity": row["severity"],
                "message": row["message"],
                "fix_action": row["fix_action"],
                "note": row["note"],})

    return rules

def get_basic_rules(check_rules: list) -> list:
    """
    基本情報欄で使うルールを取り出す関数
    """

    basic_categories = ["基本情報", "位置情報", "地況"]

    return [rule
        for rule in check_rules
        if rule["category"] in basic_categories]

def get_tree_rules(check_rules: list) -> list:
    """
    毎木調査欄で使うルールを取り出す関数
    """

    return [rule
        for rule in check_rules
        if rule["category"] == "毎木"]


# ==============================================================================================
# OCRによるPDF処理
# ==============================================================================================

def ocr_pdf(pdf_bytes, zoom: float = 2.0): # zoom=3.0でも精度改善が見られず
    """
    PDFのバイト列を受け取り、全ページの文字をテキストとして返す関数。
    pymupdfで1ページずつ画像化 → Vision APIでOCR する方式に変更。
    Pythonのモジュールで、表形式に強いためpymupdfを使用。

    zoom: 画像化するときの解像度の倍率（72dpi × zoom）。
          検証の結果、2.0 が精度・速度・メモリのバランスが良いため既定値にする。
    """

    # Google Cloud Vision（OCRサービス）と話すための"窓口"を1つ作って、clientという名前で持っておく
    # Googleのサーバー（クラウド） がOCRを行い、その結果をclientで受け取る
    # DEFAULT_GCP_KEYとos.environ.setdefaultを環境変数として認証
    client = vision.ImageAnnotatorClient() # 「電話機を1台、用意する」イメージ

    # バイト列からPDFを開く（ファイルではなくメモリ上のデータなのでstream=を使う）
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")

    # 指定なしでは72dpiで画像化するため、zoomで倍にする
    matrix = fitz.Matrix(zoom, zoom)  # 解像度の拡大率（72dpi × zoom）

    page_texts = []
    for page in doc:
        # ページを画像（pixmap）に変換し、PNGのバイト列にする
        # PDFというファイル形式」のままでは画像用のOCRに渡せないので、ページを一度PNG画像として描き直す
        pix = page.get_pixmap(matrix=matrix) 

        # pix（pymupdf独自の画像データ）をpng形式のバイト列に変換
        # バイト列を使うことで、ディスクにpngを保存せずメモリ上だけで処理
        png_bytes = pix.tobytes("png")  

        # 画像をVision APIでOCR（画像なのでdocument_text_detectionが使える）
        image = vision.Image(content=png_bytes) # contentにより、画像データそのものをGoogleサーバーに渡せる

        # 画像データをGoogleサーバーにclientを通して渡す
        # 帳票形式であるため、DOCUMENT_TEXT_DETECTION を選択
        response = client.document_text_detection(image=image) 

        # full_text_annotationがOCRした文書全体のまとまり
        page_texts.append(response.full_text_annotation.text) 

    # 全ページを区切り線でつなげて1本のテキストにする
    # ocr_pdf_endpointでは、関数の戻り値がreturn {"text": text}で1つである必要がある
    return "\n----- page break -----\n".join(page_texts) # joinは、要素の間にだけ区切りを入れる


# ==============================================================================================
# 既存エラー判別コードによるチェック
# ==============================================================================================

def calc_average(values:float)->float:
    """
    平均値を計算する関数
    空リストの場合は None を返す
    """
    
    valid_values = []

    for value in values:
        if value is not None:
            # Noneでなければ、valid_valuesに樹高、枝下高、胸高直径をそれぞれ足していく
            valid_values.append(value)

    if len(valid_values) == 0:
        return None
    
    # valid_valuesの平均値を、valid_valuesの個数で算出する
    # 樹高、枝下高、胸高直径の平均値算出に必要
    return sum(valid_values) / len(valid_values)

def make_error(sheet_name, row_no, item_name, rule):
    """
    error_log.csv に出力するエラー情報を作成する関数
    """

    return {
        "sheet_name": sheet_name,
        "row_no": row_no,
        "rule_id": rule["rule_id"],
        "category": rule["category"],
        "item_name": item_name,
        "check_item": rule["check_item"],
        "severity": rule["severity"],
        "error_message": rule["message"],
        "fix_action": rule["fix_action"],}

# 基本情報欄をチェックする
def check_basic_info(ws, sheet_name,basic_information:list,
                     basic_rules: list)-> list:
    """
    基本情報欄をcheck_rules_forest_survey.csv のルールでチェックする関数
    """

    errors = []

    basic_cell_map = dict(basic_information)

    # 未入力のエラーを抽出
    for rule in basic_rules:
        item_name = rule["target_column"]
        condition = rule["condition"]

        if item_name not in basic_cell_map:
            continue

        cell_address = basic_cell_map[item_name]
        value = ws[cell_address].value

        # 未入力チェック
        if condition == "セルが空欄、または「空白」が選択されている":
            if is_blank(value):
                errors.append(
                    make_error(sheet_name, "基本情報", item_name, rule))

        # 数値チェック
        elif condition == "数値でない":
            if not is_blank(value) and to_float(value) is None:
                errors.append(
                    make_error(sheet_name, "基本情報", item_name, rule))

        # 傾斜度 0〜90 チェック
        elif condition == "0未満、または90超":
            value_float = to_float(value)

            if value_float is not None and (value_float < 0 or value_float > 90):
                errors.append(
                    make_error(sheet_name, "基本情報", item_name, rule))

    return errors

# 毎木調査欄をチェックする
def check_tree_rows(ws, sheet_name, tree_information: dict, 
                    start_row: int, end_row: int, tree_rules: list) -> list:

    errors = []

    for row in range(start_row, end_row + 1):

        row_values = {}

        # dict型をfor文でループするために、itemsメソッドを使用する
        for item_name, col in tree_information.items():
            row_values[item_name] = ws.cell(row=row, column=col).value

        # is_unused_tree_row()を使って空行を判定する
        if is_unused_tree_row(
            row_values.get("樹種"),
            row_values.get("樹高(m)"),
            row_values.get("枝下高(m)"),
            row_values.get("胸高直径(cm)"),
            row_values.get("異常区分"),
            row_values.get("被害区分")):
            continue

        for rule in tree_rules:
            condition = rule["condition"]
            target_column = rule["target_column"]

            # target_column が「樹高_m,枝下高_m」のような場合に分割する
            target_columns = [x.strip() for x in target_column.split(",")]

            # 基本的には先頭の項目をエラー項目として扱う
            item_name = target_columns[0]
            value = row_values.get(item_name)

            # R015：樹木番号未入力
            if condition == "樹木データがある行で番号が空欄":
                if is_blank(row_values.get("番号")):
                    errors.append(make_error(sheet_name, row, "番号", rule))

            # R016：樹種未入力
            elif condition == "胸高直径,樹高,枝下高,異常区分,被害区分,備考のいずれかが入力されているのに樹種が空欄":
                has_other_data = (
                    not is_blank(row_values.get("胸高直径(cm)"))
                    or not is_blank(row_values.get("樹高(m)"))
                    or not is_blank(row_values.get("枝下高(m)"))
                    or not is_blank(row_values.get("異常区分"))
                    or not is_blank(row_values.get("被害区分"))
                    or not is_blank(row_values.get("備考")))

                if has_other_data and is_blank(row_values.get("樹種")):
                    errors.append(make_error(sheet_name, row, "樹種", rule))

            # R017：樹高空欄・枝下高あり
            elif condition == "樹高が空欄、かつ枝下高が入力されている":
                if is_blank(row_values.get("樹高(m)")) and not is_blank(row_values.get("枝下高(m)")):
                    errors.append(make_error(sheet_name, row, "樹高(m)", rule))

            # R018：樹高 0以下
            elif condition == "0以下" and item_name == "樹高(m)":
                value_float = to_float(row_values.get("樹高(m)"))
                if value_float is not None and value_float <= 0:
                    errors.append(make_error(sheet_name, row, "樹高(m)", rule))

            # R019：枝下高空欄・樹高あり
            elif condition == "枝下高が空欄、かつ樹高が入力されている":
                if is_blank(row_values.get("枝下高(m)")) and not is_blank(row_values.get("樹高(m)")):
                    errors.append(make_error(sheet_name, row, "枝下高(m)", rule))

            # R020：枝下高 0未満
            elif condition == "0未満" and item_name == "枝下高(m)":
                value_float = to_float(row_values.get("枝下高(m)"))
                if value_float is not None and value_float < 0:
                    errors.append(make_error(sheet_name, row, "枝下高(m)", rule))

            # R021：枝下高が樹高より大きい
            elif condition == "枝下高が樹高より大きい":
                branch_height = to_float(row_values.get("枝下高(m)"))
                tree_height = to_float(row_values.get("樹高(m)"))

                if branch_height is not None and tree_height is not None:
                    if branch_height > tree_height:
                        errors.append(make_error(sheet_name, row, "枝下高(m)", rule))

            # R022：枝下高が樹高と同じ
            elif condition == "枝下高が樹高と同じ値":
                branch_height = to_float(row_values.get("枝下高(m)"))
                tree_height = to_float(row_values.get("樹高(m)"))

                if branch_height is not None and tree_height is not None:
                    if branch_height == tree_height:
                        errors.append(make_error(sheet_name, row, "枝下高(m)", rule))

            # R023：胸高直径未入力
            elif condition == "樹種が入力されているのに胸高直径が空欄":
                if not is_blank(row_values.get("樹種")) and is_blank(row_values.get("胸高直径(cm)")):
                    errors.append(make_error(sheet_name, row, "胸高直径(cm)", rule))

            # R024：胸高直径 0以下
            elif condition == "0以下" and item_name == "胸高直径(cm)":
                value_float = to_float(row_values.get("胸高直径(cm)"))
                if value_float is not None and value_float <= 0:
                    errors.append(make_error(sheet_name, row, "胸高直径(cm)", rule))

            # R025：異常区分未入力
            elif condition == "樹種が入力されているのに異常区分が空欄、または「空白」が選択されている":
                if not is_blank(row_values.get("樹種")) and is_blank(row_values.get("異常区分")):
                    errors.append(make_error(sheet_name, row, "異常区分", rule))

            # R026：被害区分未入力
            elif condition == "樹種が入力されているのに被害区分が空欄、または「空白」が選択されている":
                if not is_blank(row_values.get("樹種")) and is_blank(row_values.get("被害区分")):
                    errors.append(make_error(sheet_name, row, "被害区分", rule))

    return errors


# ==============================================================================================
# LLM用：調査データをテキスト形式にまとめる
# ==============================================================================================

def extract_survey_text_for_llm(ws, sheet_name, tree_information: dict, 
                                start_row: int, end_row: int) -> str:
    """
    LLMに渡すための調査データをテキスト形式に変換する関数
    """

    lines = []
    lines.append(f"【シート名】{sheet_name}")
    lines.append(f"調査ID: {ws['B2'].value}")
    lines.append(f"地域名: {ws['D2'].value}")
    lines.append(f"調査日時: {ws['F2'].value}")
    lines.append(f"天気: {ws['H2'].value}")
    lines.append(f"記帳者: {ws['B3'].value}")
    lines.append(f"対象樹種: {ws['D3'].value}")
    lines.append(f"面積(㎡): {ws['F3'].value}")
    lines.append(f"斜面位置: {ws['H3'].value}")
    lines.append(f"斜面方位: {ws['B4'].value}")
    lines.append(f"傾斜度: {ws['D4'].value}")
    lines.append(f"緯度: {ws['F4'].value}")
    lines.append(f"経度: {ws['H4'].value}")
    lines.append("")
    lines.append("【毎木調査データ】")

    for row in range(start_row, end_row + 1):
        row_values = {}
        for item_name, col in tree_information.items():
            row_values[item_name] = ws.cell(row=row, column=col).value

        # 空行はスキップ
        if is_unused_tree_row(
            row_values.get("樹種"),
            row_values.get("樹高(m)"),
            row_values.get("枝下高(m)"),
            row_values.get("胸高直径(cm)"),
            row_values.get("異常区分"),
            row_values.get("被害区分")):
            continue

        # 各行を1行のテキストにまとめる
        row_text = f"行{row}: " + " / ".join(
            f"{k}={v}" for k, v in row_values.items() if not is_blank(v))
        lines.append(row_text)

    return "\n".join(lines)


# ==============================================================================================
# GIS結合用csv出力
# ==============================================================================================

def create_gis_summary_row(ws, sheet_name, tree_information: dict, 
                           start_row: int, end_row:int)-> dict:
    """
    GIS結合用CSVに出力する1地点分の集計行を作成する
    """

    # 基本情報
    plot_id = ws["B2"].value # 調査IDのセル値取得
    area_name = ws["D2"].value # 地域名のセル値取得
    survey_date = ws["F2"].value # 調査日時のセル値取得
    weather = ws["H2"].value # 天気のセル値取得
    writer = ws["B3"].value # 記帳者のセル値取得
    tree_species = ws["D3"].value # 対象樹種のセル値取得
    area = ws["F3"].value # 面積(㎡)ののセル値取得
    slope_position = ws["H3"].value # 斜面位置のセル値取得
    slope_aspect = ws["B4"].value # 斜面方位のセル値取得
    inclination = ws["D4"].value # 傾斜度のセル値取得
    latitude = ws["F4"].value # 緯度のセル値取得
    longitude = ws["H4"].value # 経度のセル値取得
    
    tree_heights = []
    branch_heights = []
    dbhs = []
    tree_count = 0

    for row in range(start_row, end_row + 1):
        tree_species_value = ws.cell(row=row, column=tree_information["樹種"]).value

        # 樹種が空欄なら、その行は立木データとして扱わない
        if is_blank(tree_species_value):
            continue

        tree_count += 1

        tree_height = to_float(ws.cell(row=row, column=tree_information["樹高(m)"]).value)
        branch_height = to_float(ws.cell(row=row, column=tree_information["枝下高(m)"]).value)
        dbh = to_float(ws.cell(row=row, column=tree_information["胸高直径(cm)"]).value)

        if tree_height is not None:
            tree_heights.append(tree_height)

        if branch_height is not None:
            branch_heights.append(branch_height)

        if dbh is not None:
            dbhs.append(dbh)

    summary_row = {
        "調査ID": plot_id,
        "シート名": sheet_name,
        "地域名": area_name,
        "調査日時": format_survey_datetime(survey_date), # 調査日時を表示用の文字列に整える
        "天気": weather,
        "記帳者": writer,
        "対象樹種": tree_species,
        "面積(㎡)": area,
        "斜面位置": slope_position,
        "斜面方位": slope_aspect,
        "傾斜度": format_decimal(inclination,1), # 小数点の桁数を定義
        "緯度": format_decimal(latitude,6), # 小数点の桁数を定義
        "経度": format_decimal(longitude,6), # 小数点の桁数を定義
        "平均樹高(m)": format_decimal(calc_average(tree_heights),1), # 小数点の桁数を定義
        "平均枝下高(m)": format_decimal(calc_average(branch_heights),1), # 小数点の桁数を定義
        "平均胸高直径(cm)": format_decimal(calc_average(dbhs),1), # 小数点の桁数を定義
        "立木本数": tree_count,}

    return summary_row

def write_gis_csv(gis_rows, output_path:str)-> list:
    """
    QGIS結合用CSVを出力する関数
    """

    fieldnames = ["調査ID","シート名","地域名","調査日時","天気",
        "記帳者","対象樹種","面積(㎡)","斜面位置",
        "斜面方位","傾斜度","緯度","経度",
        "平均樹高(m)","平均枝下高(m)","平均胸高直径(cm)","立木本数",]

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(gis_rows)


# ==============================================================================================
# Google Cloud Vision API呼び出し
# ==============================================================================================

@app.post("/ocr_pdf")
async def ocr_pdf_endpoint(
    pdf_file: UploadFile = File(...)):
    """
    PDFを受け取りOCRでテキストを抽出するエンドポイント
    """

    # awaitを付けることで、非同期処理
    pdf_bytes = await pdf_file.read()
    text = ocr_pdf(pdf_bytes)

    # ページ別の文字数を確認（空ページや極端に少ない場合は警告）
    # Dify上で、OCRの読み取り失敗か本当に空欄が判別できなかったため追加
    pages = text.split("\n----- page break -----\n") # ページの区切り

    warnings = []

    for i, page_text in enumerate(pages):
        if len(page_text.strip()) < 50:  # 50文字未満は怪しい(経験則)
            # あまり少ない文字数だとwaringsの数が増加する
            warnings.append(f"ページ{i+1}: 文字数が少ない（{len(page_text.strip())}文字）")
    # Dify側で、文字数の警告確認
    return {"text": text,
            "ocr_warnings": warnings}

# DifyのLLM(OCRからの文字抽出)で使用
def is_ocr_unreadable(value: str) -> bool:
    """
    OCRで読み取れなかった項目かどうか判定
    OCRで読み取れなかったものは、"OCR_UNREADABLE"で出力
    """

    return str(value).strip() == "OCR_UNREADABLE"


# ==============================================================================================
# エラーあり
# エラーログをUTF-8 BOM付きCSVで返す
# DifyのMarkDownエクスポーターでは文字化けが解消できなかったため追加
# ==============================================================================================

@app.post("/export_error_log")
async def export_error_log(
    survey_file: UploadFile = File(...)):
    """
    エラーログをUTF-8 BOM付きCSVファイルとして返すエンドポイント
    Excelで直接開いても文字化けしない
    """
    
    # awaitを付けることで、非同期処理
    survey_bytes = await survey_file.read()
    wb = load_workbook(io.BytesIO(survey_bytes), data_only=True)

    cell_mapping_file = MASTER_DIR / "forest_survey_cell_mapping.xlsx"
    check_rules_file = MASTER_DIR / "check_rules_forest_survey.csv"

    basic_information = load_basic_inf_cell_mapping(cell_mapping_file)
    tree_information = dict(load_tree_inf_cell_mapping(cell_mapping_file))
    start_row, end_row = load_tree_data_cell_mapping(cell_mapping_file)

    check_rules = load_check_rules(check_rules_file)
    basic_rules = get_basic_rules(check_rules)
    tree_rules = get_tree_rules(check_rules)

    all_errors = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_errors.extend(check_basic_info(ws, sheet_name, basic_information, basic_rules))
        all_errors.extend(check_tree_rows(ws, sheet_name, tree_information, 
                                          start_row, end_row, tree_rules))
        
    fieldnames = [
        "sheet_name", "row_no", "rule_id", "category",
        "item_name", "check_item", "severity",
        "error_message", "fix_action",]

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(all_errors)
    
    # BOM(Byte Order Mark)を付けることで、 UTF-8形式として認識する
    encoded = output.getvalue().encode("utf-8-sig") 

    return Response(
        content=encoded,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=error_log.csv"})


# ==============================================================================================
# 追加エラー候補追加
# エラー判別コードに新たなルールを追加する
# ==============================================================================================

# 適用できていない(今後の課題)
@app.post("/add_rules")
async def add_rules(request: Request):
    """
    LLMが発見した新しいルールを check_rules_forest_survey.csv に追記する
    リスト形式・JSON文字列形式どちらでも受け付ける
    """

    # bodyをそのまま受け取り、柔軟にパース(構造化処理)する
    raw_body = await request.body()

    try:
        parsed = json.loads(raw_body)
        # JSON文字列がさらにネストされている場合（"[...]" のような文字列）も対応
        if isinstance(parsed, str):
            parsed = json.loads(parsed)
        if not isinstance(parsed, list):
            return {"added_count": 0, "status": "invalid_format"}
        new_rules = parsed
    except Exception:
        return {"added_count": 0, "status": "parse_error"}

    # ルールが空の場合はそのまま返す
    if not new_rules:
        return {"added_count": 0, "status": "ok"}

    check_rules_file = MASTER_DIR / "check_rules_forest_survey.csv"

    # 既存のルールを読み込んで、既存のrule_idを取得する
    existing_rules = load_check_rules(check_rules_file)
    existing_ids = [r["rule_id"] for r in existing_rules]

    # 追記用にファイルを開く
    fieldnames = [
        "rule_id", "category", "check_item",
        "target_column", "condition", "severity",
        "message", "fix_action", "note"]

    added_count = 0
    with open(check_rules_file, "a", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")

        for rule in new_rules:
            # すでに同じrule_idがある場合はスキップ
            if rule.get("rule_id") in existing_ids:
                continue

            # LLMが出力するフィールド名を正規フィールド名に変換する
            normalized = {
                "rule_id":       rule.get("rule_id", ""),
                "category":      rule.get("category", ""),
                "check_item":    rule.get("check_item", ""),
                "target_column": rule.get("target_column") or rule.get("item_name", ""),
                "condition":     rule.get("condition", ""),
                "severity":      rule.get("severity", "warning"),
                "message":       rule.get("message") or rule.get("error_message", ""),
                "fix_action":    rule.get("fix_action", ""),
                "note":          rule.get("note", ""),}
            writer.writerow(normalized)
            added_count += 1

    return {"added_count": added_count, "status": "ok"}


# ==============================================================================================
# エラーなし(GIS用フォーマット出力)
# DifyのMarkDownエクスポーターでは文字化けが解消できなかったため、UTF-8 BOM付きCSVで返す処理を追加
# ==============================================================================================

@app.get("/")
def read_root():
    return {"status": "ok"}

# GIS用サマリーをUTF-8 BOM付きCSVで返す
@app.post("/export_gis_csv")
async def export_gis_csv(
    survey_file: UploadFile = File(...)):
    """
    GIS結合用サマリーをUTF-8 BOM付きCSVファイルとして返すエンドポイント
    Excelで直接開いても文字化けしない
    エラーがある場合は空のCSVを返す
    """

     # awaitを付けることで、非同期処理
    survey_bytes = await survey_file.read()
    # メモリ上でバイナリデータをファイルのように扱う
    # (ディスクに書き込まずに、メモリ上で「仮想的なファイル」として扱える)
    wb = load_workbook(io.BytesIO(survey_bytes), data_only=True)

    cell_mapping_file = MASTER_DIR / "forest_survey_cell_mapping.xlsx"
    check_rules_file = MASTER_DIR / "check_rules_forest_survey.csv"

    basic_information = load_basic_inf_cell_mapping(cell_mapping_file)
    tree_information = dict(load_tree_inf_cell_mapping(cell_mapping_file))
    start_row, end_row = load_tree_data_cell_mapping(cell_mapping_file)

    check_rules = load_check_rules(check_rules_file)
    basic_rules = get_basic_rules(check_rules)
    tree_rules = get_tree_rules(check_rules)

    all_errors = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_errors.extend(check_basic_info(ws, sheet_name, basic_information, basic_rules))
        all_errors.extend(check_tree_rows(ws, sheet_name, tree_information, start_row, 
                                          end_row, tree_rules))

    fieldnames = ["調査ID", "シート名", "地域名", "調査日時", "天気",
        "記帳者", "対象樹種", "面積(㎡)", "斜面位置", 
        "斜面方位","傾斜度", "緯度", "経度",
        "平均樹高(m)", "平均枝下高(m)", "平均胸高直径(cm)", "立木本数",]

    gis_rows = []
    
    if not all_errors:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            gis_rows.append(
                create_gis_summary_row(ws, sheet_name, tree_information, start_row, end_row))

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(gis_rows)

    encoded = output.getvalue().encode("utf-8-sig")

    return Response(
        content=encoded,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=gis_plot_summary.csv"})


# ==============================================================================================
# メインの処理
# ==============================================================================================

@app.post("/analyze")
async def analyze(
    survey_file: UploadFile = File(...)):

    # 1．架空の森林調査票をアップロードする
    # ネットワークでアップロードされたファイルは、バイト列になる
    # ファイルの読み込みでは、非同期処理が必要
    # (awaitは、データが送付されるのを待ってから次の処理に進んでくださいという指示)
    survey_bytes = await survey_file.read() 

    # バイト列をopenpyxlで読み込む(load_workbookuだけでは、
    # バイト列をそのまま渡すことになるためio.BytesIOを使用する)
    wb = load_workbook(io.BytesIO(survey_bytes), data_only=True)

    # 2. masterフォルダから固定ファイルを読み込む
    cell_mapping_file = MASTER_DIR / "forest_survey_cell_mapping.xlsx"
    check_rules_file = MASTER_DIR / "check_rules_forest_survey.csv"

    # 3. セル対応表を読み込む
    basic_information = load_basic_inf_cell_mapping(cell_mapping_file)
    tree_information = dict(load_tree_inf_cell_mapping(cell_mapping_file))
    start_row, end_row = load_tree_data_cell_mapping(cell_mapping_file)

    # 4. ルールを読み込む
    check_rules = load_check_rules(check_rules_file)
    basic_rules = get_basic_rules(check_rules)
    tree_rules = get_tree_rules(check_rules)

    # 5. シートごとにエラーチェックを実行する
    target_sheets = wb.sheetnames

    all_errors = []

    for sheet_name in target_sheets:
        ws = wb[sheet_name]

        basic_errors = check_basic_info(ws, sheet_name, basic_information, basic_rules,)

        tree_errors = check_tree_rows(ws, sheet_name, tree_information, start_row, end_row, tree_rules,)

        all_errors.extend(basic_errors)
        all_errors.extend(tree_errors)
    
    # 6. エラーがない場合のみGISサマリーを作成する
    gis_rows = []

    if not all_errors:
        for sheet_name in target_sheets:
            ws = wb[sheet_name]
            summary_row = create_gis_summary_row(ws, sheet_name, tree_information, start_row, end_row)
            gis_rows.append(summary_row)

    # 7. LLM用のテキストデータを生成する(追加エラー候補抽出用)
    survey_texts = []
    for sheet_name in target_sheets:
        ws = wb[sheet_name]
        text = extract_survey_text_for_llm(ws, sheet_name, tree_information, start_row, end_row)
        survey_texts.append(text)

    survey_text_for_llm = "\n\n".join(survey_texts)

    # 8. 結果を返す
    return {
        "has_error": len(all_errors) > 0,
        "error_count": len(all_errors),
        "error_log": all_errors,
        "gis_summary": gis_rows,
        "survey_text_for_llm": survey_text_for_llm}  # Dify LLMに渡す用


# ==============================================================================================
# PDF/OCR → Excelフォーマット復元
# ==============================================================================================

def build_survey_workbook(plots: list, template_file, 
                          cell_mapping_file) -> io.BytesIO:
    """
    構造化済みのJSON(plots)を、調査票テンプレートに流し込んで
    1地点=1シートのxlsx（バイト列）を作る関数。

    plots の形（1要素=1地点）:
        {
          "調査ID": "P001", "地域名": "A地区", ... "経度": 134.14,
          "trees": [
            {"番号":1,"樹種":"スギ","樹高(m)":18.5,"枝下高(m)":10.6,
             "胸高直径(cm)":42.8,"異常区分":"なし","被害区分":"なし","備考":""},
            ...
          ]
        }
    """

    # セル対応表から「項目名→セル位置」「項目名→列番号」「データ行範囲」を取得する
    # ＝既存のチェック処理と同じ対応表を使うので、出力フォーマットが完全に揃う
    basic_information = load_basic_inf_cell_mapping(cell_mapping_file)      # [(項目名, "B2"), ...]
    tree_information = dict(load_tree_inf_cell_mapping(cell_mapping_file))  # {"樹種": 2, ...}
    start_row, end_row = load_tree_data_cell_mapping(cell_mapping_file)     # 7, 76

    # テンプレートを開く（data_only=Falseで集計セルの数式を保持する）
    wb = load_workbook(template_file)
    template_ws = wb[wb.sheetnames[0]]

    # 地点が無ければテンプレートをそのまま返す（空シート保持のため）
    if not plots:
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
    
    # シート名の重複を避けるためsetを使う
    used_names = set() 

    for i, plot in enumerate(plots):
        # シート名は調査ID優先。無ければPlot連番。
        raw_name = str(plot.get("調査ID") or f"Plot{i + 1}").strip()
        # Excelのシート名は31文字まで
        sheet_name = (raw_name or f"Plot{i + 1}")[:31] 
        # 同名シートを避ける
        base = sheet_name
        n = 2
        while sheet_name in used_names:
            sheet_name = f"{base[:28]}_{n}"
            n += 1
        used_names.add(sheet_name)

        # テンプレートをコピーして新しいシートを作る（書式・数式ごと複製）
        ws = wb.copy_worksheet(template_ws)
        ws.title = sheet_name

        # 基本情報を所定のセルへ書き込む（例: 調査ID→B2）
        for item_name, cell_address in basic_information:
            if item_name in plot and plot[item_name] is not None:
                ws[cell_address] = plot[item_name]

        # 毎木データを7行目から順に書き込む
        trees = plot.get("trees", []) or []
        for idx, tree in enumerate(trees):
            row = start_row + idx
            if row > end_row:
                break  # 最大70本を超えた分は捨てる
            for item_name, col in tree_information.items():
                if item_name in tree and tree[item_name] not in (None, ""):
                    ws.cell(row=row, column=col, value=tree[item_name])

    # 元のテンプレートシートは不要なので削除する
    wb.remove(template_ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# xlsxを返すときのMIMEタイプ
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ==============================================================================================
# PDF/OCRで抽出した調査データをCSVとして出力
# ==============================================================================================

@app.post("/export_ocr_csv")
async def export_ocr_csv(request: Request):
    """
    DifyのLLMノードで構造化したJSONを受け取り、
    OCRで抽出した調査結果をUTF-8 BOM付きCSVとして返す。
    出力形式:
      1行 = 1本の立木データ。
      地点の基本情報は、立木データの各行へ繰り返して記録する。
      立木データが無い地点も、基本情報だけで1行出力する。
    """

    # Difyから送信されたJSONを受け取る
    # 非同期処理
    raw_body = await request.body()

    try:
        # JSONをPythonで扱えるdict型またはlist型に変換する
        data = json.loads(raw_body)
        # dataの方が文字列かどうか判定する
        # JSON文字列が二重に入れ子になっている場合にも対応する
        if isinstance(data, str):
            data = json.loads(data)
    
    # 全ての例外を対象とする
    # 構造化データ（辞書やリスト、オブジェクトなど）に変換する処理(パース)に失敗してもエラーにならない
    except Exception:
        return {"status": "parse_error"}

    # {"plots": [...]} と [...] の両方に対応する
    # dataのオブジェクト型がdictかlistかそれ以外かどうか判定する
    # dataのオブジェクト型がdictかlistでもなければ空にする
    if isinstance(data, dict): 
        plots = data.get("plots", [])
    elif isinstance(data, list):
        plots = data
    else:
        plots = []

    # CSVの列名
    fieldnames = ["調査ID","地域名","調査日時","天気",
        "記帳者","対象樹種","面積(㎡)","斜面位置",
        "斜面方位","傾斜度","緯度","経度",
        "番号","樹種","樹高(m)","枝下高(m)","胸高直径(cm)","異常区分","被害区分","備考",]

    # 地点全体に関する項目
    basic_fields = ["調査ID","地域名","調査日時","天気",
        "記帳者","対象樹種","面積(㎡)","斜面位置",
        "斜面方位","傾斜度","緯度","経度",]

    # 立木1本ごとに異なる項目
    tree_fields = ["番号","樹種","樹高(m)","枝下高(m)","胸高直径(cm)",
                   "異常区分","被害区分","備考",]

    rows = []

    for plot in plots:
        if not isinstance(plot, dict):
            continue

        # 地点の基本情報を取得する
        basic_row = {
            field: plot.get(field, "")
            for field in basic_fields}

        trees = plot.get("trees", []) or []

        # 立木データが無い地点も、地点情報だけで1行残す
        if not trees:
            row = basic_row.copy()

            row.update({field: ""
                for field in tree_fields})

            rows.append(row)
            continue

        # 立木1本につきCSVを1行作成する
        for tree in trees:
            # treeがdict型でなければforから抜け出す
            if not isinstance(tree, dict):
                continue

            row = basic_row.copy()

            row.update({field: tree.get(field, "")
                for field in tree_fields})

            rows.append(row)

    # CSVをメモリ上で作成する
    output = io.StringIO()

    writer = csv.DictWriter(output,fieldnames=fieldnames)

    writer.writeheader()
    writer.writerows(rows)

    # Excelで文字化けしにくい形式にする
    encoded = output.getvalue().encode("utf-8-sig")

    return Response(
        content=encoded,
        media_type="text/csv",
        headers={"Content-Disposition":"attachment; filename=ocr_survey_results.csv"},)

# Dify側でLLM構造化 → そのJSONを受け取りxlsxを返す
@app.post("/build_xlsx_from_json")
async def build_xlsx_from_json(request: Request):
    """
    構造化済みJSON（{"plots":[...]} もしくは [...] 形式）を受け取り、
    調査票テンプレートに流し込んだxlsxを返すエンドポイント。
    DifyのLLMノードで作ったJSONをここへPOSTする想定（パターンA）。
    """

    # 非同期処理
    raw_body = await request.body()
    
    # 例外処理
    try: # json文字列をPythonオブジェクトに変換
        data = json.loads(raw_body)

        # dataの方が文字列であるかどうか判定
        # 文字列が二重にネスト(入れ子)の場合は、データにアクセスできないためでコードする
        # 2回json.loads()使うことで、ネストをほどきアクセスできるようにする
        # LLMの出力では、JSON文字列がさらに文字列としてネストされていることがあるため必要
        if isinstance(data, str):
            data = json.loads(data)

    # 全ての例外を対象とする
    # 構造化データ（辞書やリスト、オブジェクトなど）に変換する処理(パース)に失敗してもエラーにならないようにする
    except Exception:
        return {"status": "parse_error"}

    # {"plots":[...]} でも [...] でも受け付ける
    # dataの型がdicyかlistかどうか判定
    if isinstance(data, dict):
        # dict型から"plots"のキーを取得する、なければ空のlistを取得する
        # getメソッドを使うことで、キーがなくてもエラーを発生させずに任意の値を取得できる
        plots = data.get("plots", [])
    elif isinstance(data, list):
        plots = data
    else:
        plots = []

    template_file = MASTER_DIR / "template_forest_survey.xlsx"
    cell_mapping_file = MASTER_DIR / "forest_survey_cell_mapping.xlsx"

    buf = build_survey_workbook(plots, template_file, cell_mapping_file)

    return Response(
        content=buf.getvalue(), # バイト列を取り出す
        media_type=XLSX_MEDIA_TYPE, # Excelの MIMEタイプ
        headers={"Content-Disposition": "attachment; filename=survey_from_pdf.xlsx"},)