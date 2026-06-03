from fastapi import FastAPI, UploadFile, File
from openpyxl import load_workbook
from pathlib import Path
import io
import csv

app = FastAPI()

# masterフォルダのパスを定義する
MASTER_DIR = Path(__file__).parent / "master"

# 1. ファイルの読み込み
# 空欄、Noneを判定する
def is_blank(value: str) -> bool:
    """
    空欄判定
    None、空文字、スペースのみを空欄とみなす
    """
    if value is None:
        return True
    # 半角・全角スペースの両方を除去して判定する
    cleaned = str(value).strip().replace("\u3000", "")  # ← == "" を削除
    return cleaned == ""  # ← ここで比較する

def to_float(value:float)-> float:
    """
    数値に変換する関数
    空欄または数値変換できない値は None を返す
    """
    # 数値を一度浮動小数点型に変換

    if is_blank(value):
        return None

    try:
        return float(value)
    except ValueError:
        return None

def format_decimal(value, digits:float)-> float:
    """
    小数点以下の桁数をそろえる関数
    """
    value_float = to_float(value)

    if value_float is None:
        return ""

    return f"{value_float:.{digits}f}"

def format_survey_datetime(value:str)-> str:
    """
    調査日時を表示用の文字列に整える関数
    Excelの値をそのまま使い、空欄の場合は空文字を返す
    """
    
    if is_blank(value):
        return ""

    return str(value).strip()

def is_unused_tree_row(tree_species:str, tree_height:float, 
    branch_height:float, dbh:float, abnormal_type:str, damage_type:str)-> str:
    """
    毎木調査の未使用行かどうかを判定する関数
    番号は最初から入力されている可能性があるため、判定に使わない
    """
    return (
        is_blank(tree_species)
        and is_blank(tree_height)
        and is_blank(branch_height)
        and is_blank(dbh)
        and is_blank(abnormal_type)
        and is_blank(damage_type))

def load_basic_inf_cell_mapping(cell_mapping_file:str)-> str:
    """
    forest_survey_cell_mapping.xlsx から
    基本情報の項目名とセル位置を取得する関数
    """

    wb = load_workbook(cell_mapping_file, data_only=True)
    sheet = wb["基本情報セル対応表"]

    basic_information = []

    for row in sheet.iter_rows(
        min_row=3,
        max_row=14,
        min_col=2,
        max_col=4,
        values_only=True):
        item_name = row[0]      # B列：項目名
        cell_address = row[2]   # D列：セル位置

        if is_blank(item_name) or is_blank(cell_address):
            continue

        basic_information.append((item_name, cell_address))

    return basic_information

def load_tree_inf_cell_mapping(cell_mapping_file:str)-> str:
    """
    forest_survey_cell_mapping.xlsx から
    毎木情報の項目名とセル位置を取得する関数
    """

    wb = load_workbook(cell_mapping_file, data_only=True)
    sheet = wb["毎木調査欄セル対応表"]
    
    tree_information = []
    i = 1

    for row in sheet.iter_rows(
        min_row=2,
        max_row=9,
        min_col=2,
        max_col=4,
        values_only=True):
        columns = row[0] # B列：項目名
        columns_num = i
        i = i + 1
       
        if is_blank(columns) or is_blank(columns_num):
            continue

        tree_information.append((columns, columns_num))

    return tree_information

def load_tree_data_cell_mapping(cell_mapping_file:str)-> tuple:
    """
    forest_survey_cell_mapping.xlsx から
    毎木情報の調査結果の開始行と終了行を取得する関数
    """

    wb = load_workbook(cell_mapping_file, data_only=True)
    sheet = wb["毎木調査欄セル対応表"]
    
    for row in sheet.iter_rows(
        min_row=2,
        max_row=9,
        min_col=2,
        max_col=4,
        values_only=True):

        cell_address = row[2]   # D列：セル範囲 例 B7:B76
        
        if is_blank(cell_address):
            continue

        start_cell = cell_address.split(":")[0]  # 例 B7
        end_cell = cell_address.split(":")[1]    # 例 B76

        start_row = int(start_cell[1:])  # 7
        end_row = int(end_cell[1:])      # 76

        return start_row, end_row

    return None, None

# check_rules_forest_survey.csvに基づいてエラーを表示する
def load_check_rules(check_rules_file: str) -> list:
    """
    check_rules_forest_survey.csv を読み込む関数
    """

    rules = []

    with open(check_rules_file, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)

        for row in reader:
            rules.append({
                "rule_id": row["rule_id"],
                "category": row["category"],
                "check_item": row["check_item"],
                "target_column": row["target_column"],
                "condition": row["condition"],
                "severity": row["severity"],
                "message": row["message"],
                "fix_action": row["fix_action"],
                "note": row["note"],
            })

    return rules

def get_basic_rules(check_rules: list) -> list:
    """
    基本情報欄で使うルールを取り出す関数
    """

    basic_categories = ["基本情報", "位置情報", "地況"]

    return [
        rule
        for rule in check_rules
        if rule["category"] in basic_categories]

def get_tree_rules(check_rules: list) -> list:
    """
    毎木調査欄で使うルールを取り出す関数
    """

    return [
        rule
        for rule in check_rules
        if rule["category"] == "毎木"]


# 2. エラーチェック
def calc_average(values:float)->float:
    """
    平均値を計算する関数
    空リストの場合は None を返す
    """
    # 樹高、枝下高、胸高直径の平均値算出に必要

    valid_values = []

    for value in values:
        if value is not None:
            valid_values.append(value)

    if len(valid_values) == 0:
        return None
  
    return sum(valid_values) / len(valid_values)

def make_error(sheet_name, row_no, item_name, rule):
    """
    error_log.csv に出力するエラー情報を作成する関数
    """

    return {
        "sheet_name": sheet_name,
        "row_no": row_no,
        "rule_id": rule["rule_id"],
        "category": rule["category"],
        "item_name": item_name,
        "check_item": rule["check_item"],
        "severity": rule["severity"],
        "error_message": rule["message"],
        "fix_action": rule["fix_action"],}

# 基本情報欄をチェックする
def check_basic_info(ws, sheet_name,basic_information:list,basic_rules: list)-> list:
    """
    基本情報欄を、check_rules_forest_survey.csv のルールでチェックする関数
    """

    errors = []

    basic_cell_map = dict(basic_information)

    # 未入力のエラーを抽出
    for rule in basic_rules:
        item_name = rule["target_column"]
        condition = rule["condition"]

        if item_name not in basic_cell_map:
            continue

        cell_address = basic_cell_map[item_name]
        value = ws[cell_address].value

        # 未入力チェック
        if condition == "セルが空欄、または「空白」が選択されている":
            if is_blank(value):
                errors.append(
                    make_error(sheet_name, "基本情報", item_name, rule))

        # 数値チェック
        elif condition == "数値でない":
            if not is_blank(value) and to_float(value) is None:
                errors.append(
                    make_error(sheet_name, "基本情報", item_name, rule))

        # 傾斜度 0〜90 チェック
        elif condition == "0未満、または90超":
            value_float = to_float(value)

            if value_float is not None and (value_float < 0 or value_float > 90):
                errors.append(
                    make_error(sheet_name, "基本情報", item_name, rule)
                )

    return errors

# 毎木調査欄をチェックする
def check_tree_rows(ws, sheet_name, tree_information: dict, start_row: int, end_row: int, tree_rules: list) -> list:

    errors = []

    for row in range(start_row, end_row + 1):

        row_values = {}

        for item_name, col in tree_information.items():
            row_values[item_name] = ws.cell(row=row, column=col).value

        # ★ is_unused_tree_row()を使って空行を判定する
        if is_unused_tree_row(
            row_values.get("樹種"),
            row_values.get("樹高(m)"),
            row_values.get("枝下高(m)"),
            row_values.get("胸高直径(cm)"),
            row_values.get("異常区分"),
            row_values.get("被害区分")):
            continue

        for rule in tree_rules:
            condition = rule["condition"]
            target_column = rule["target_column"]

            # target_column が「樹高_m,枝下高_m」のような場合に分割する
            target_columns = [x.strip() for x in target_column.split(",")]

            # 基本的には先頭の項目をエラー項目として扱う
            item_name = target_columns[0]
            value = row_values.get(item_name)

            # R015：樹木番号未入力
            if condition == "樹木データがある行で番号が空欄":
                if is_blank(row_values.get("番号")):
                    errors.append(make_error(sheet_name, row, "番号", rule))

            # R016：樹種未入力
            elif condition == "胸高直径,樹高,枝下高,異常区分,被害区分,備考のいずれかが入力されているのに樹種が空欄":
                has_other_data = (
                    not is_blank(row_values.get("胸高直径(cm)"))
                    or not is_blank(row_values.get("樹高(m)"))
                    or not is_blank(row_values.get("枝下高(m)"))
                    or not is_blank(row_values.get("異常区分"))
                    or not is_blank(row_values.get("被害区分"))
                    or not is_blank(row_values.get("備考"))
                )

                if has_other_data and is_blank(row_values.get("樹種")):
                    errors.append(make_error(sheet_name, row, "樹種", rule))

            # R017：樹高空欄・枝下高あり
            elif condition == "樹高が空欄、かつ枝下高が入力されている":
                if is_blank(row_values.get("樹高(m)")) and not is_blank(row_values.get("枝下高(m)")):
                    errors.append(make_error(sheet_name, row, "樹高(m)", rule))

            # R018：樹高 0以下
            elif condition == "0以下" and item_name == "樹高(m)":
                value_float = to_float(row_values.get("樹高(m)"))
                if value_float is not None and value_float <= 0:
                    errors.append(make_error(sheet_name, row, "樹高(m)", rule))

            # R019：枝下高空欄・樹高あり
            elif condition == "枝下高が空欄、かつ樹高が入力されている":
                if is_blank(row_values.get("枝下高(m)")) and not is_blank(row_values.get("樹高(m)")):
                    errors.append(make_error(sheet_name, row, "枝下高(m)", rule))

            # R020：枝下高 0未満
            elif condition == "0未満" and item_name == "枝下高(m)":
                value_float = to_float(row_values.get("枝下高(m)"))
                if value_float is not None and value_float < 0:
                    errors.append(make_error(sheet_name, row, "枝下高(m)", rule))

            # R021：枝下高が樹高より大きい
            elif condition == "枝下高が樹高より大きい":
                branch_height = to_float(row_values.get("枝下高(m)"))
                tree_height = to_float(row_values.get("樹高(m)"))

                if branch_height is not None and tree_height is not None:
                    if branch_height > tree_height:
                        errors.append(make_error(sheet_name, row, "枝下高(m)", rule))

            # R022：枝下高が樹高と同じ
            elif condition == "枝下高が樹高と同じ値":
                branch_height = to_float(row_values.get("枝下高(m)"))
                tree_height = to_float(row_values.get("樹高(m)"))

                if branch_height is not None and tree_height is not None:
                    if branch_height == tree_height:
                        errors.append(make_error(sheet_name, row, "枝下高(m)", rule))

            # R023：胸高直径未入力
            elif condition == "樹種が入力されているのに胸高直径が空欄":
                if not is_blank(row_values.get("樹種")) and is_blank(row_values.get("胸高直径(cm)")):
                    errors.append(make_error(sheet_name, row, "胸高直径(cm)", rule))

            # R024：胸高直径 0以下
            elif condition == "0以下" and item_name == "胸高直径(cm)":
                value_float = to_float(row_values.get("胸高直径(cm)"))
                if value_float is not None and value_float <= 0:
                    errors.append(make_error(sheet_name, row, "胸高直径(cm)", rule))

            # R025：異常区分未入力
            elif condition == "樹種が入力されているのに異常区分が空欄、または「空白」が選択されている":
                if not is_blank(row_values.get("樹種")) and is_blank(row_values.get("異常区分")):
                    errors.append(make_error(sheet_name, row, "異常区分", rule))

            # R026：被害区分未入力
            elif condition == "樹種が入力されているのに被害区分が空欄、または「空白」が選択されている":
                if not is_blank(row_values.get("樹種")) and is_blank(row_values.get("被害区分")):
                    errors.append(make_error(sheet_name, row, "被害区分", rule))

    return errors


# GIS結合用csv出力
def create_gis_summary_row(ws, sheet_name, tree_information: dict, start_row: int)-> dict:
    """
    GIS結合用CSVに出力する1地点分の集計行を作成する
    """

    # 基本情報
    plot_id = ws["B2"].value # 調査ID
    area_name = ws["D2"].value # 地域名
    survey_date = ws["F2"].value # 調査日時
    weather = ws["H2"].value # 天気
    writer = ws["B3"].value # 記帳者
    tree_species = ws["D3"].value # 対象樹種
    area = ws["F3"].value # 面積_㎡
    slope_position = ws["H3"].value # 斜面位置
    slope_aspect = ws["B4"].value # 斜面方位
    inclination = ws["D4"].value # 傾斜度
    latitude = ws["F4"].value # 緯度
    longitude = ws["H4"].value # 経度
    
    tree_heights = []
    branch_heights = []
    dbhs = []
    tree_count = 0

    for row in range(start_row, ws.max_row + 1):
        tree_species_value = ws.cell(row=row, column=tree_information["樹種"]).value

        # 樹種が空欄なら、その行は立木データとして扱わない
        if is_blank(tree_species_value):
            continue

        tree_count += 1

        tree_height = to_float(ws.cell(row=row, column=tree_information["樹高(m)"]).value)
        branch_height = to_float(ws.cell(row=row, column=tree_information["枝下高(m)"]).value)
        dbh = to_float(ws.cell(row=row, column=tree_information["胸高直径(cm)"]).value)

        if tree_height is not None:
            tree_heights.append(tree_height)

        if branch_height is not None:
            branch_heights.append(branch_height)

        if dbh is not None:
            dbhs.append(dbh)

    summary_row = {
        "調査ID": plot_id,
        "シート名": sheet_name,
        "地域名": area_name,
        "調査日時": format_survey_datetime(survey_date),
        "天気": weather,
        "記帳者": writer,
        "対象樹種": tree_species,
        "面積(㎡)": area,
        "斜面位置": slope_position,
        "斜面方位": slope_aspect,
        "傾斜度": format_decimal(inclination,1),
        "緯度": format_decimal(latitude,6),
        "経度": format_decimal(longitude,6),
        
        "平均樹高(m)": format_decimal(calc_average(tree_heights),1),
        "平均枝下高(m)": format_decimal(calc_average(branch_heights),1),
        "平均胸高直径(cm)": format_decimal(calc_average(dbhs),1),
        "立木本数": tree_count,
    }

    return summary_row

def write_gis_csv(gis_rows, output_path:str)-> list:
    """
    QGIS結合用CSVを出力する関数
    """

    fieldnames = [
        "調査ID",
        "シート名",
        "地域名",
        "調査日時",
        "天気",
        "記帳者",
        "対象樹種",
        "面積(㎡)",
        "斜面位置",
        "斜面方位",
        "傾斜度",
        "緯度",
        "経度",
        "平均樹高(m)",
        "平均枝下高(m)",
        "平均胸高直径(cm)",
        "立木本数",]

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(gis_rows)


@app.get("/")
def read_root():
    return {"status": "ok"}

@app.post("/analyze")
async def analyze(
    survey_file: UploadFile = File(...)
):  
    # 1．架空の森林調査票をアップロードする
    # ネットワークでアップロードされたファイルは、バイト列になる
    # ファイルの読み込みでは、非同期処理が必要
    # (awaitは、データが送付されるのを待ってから次の処理に進んでくださいという指示)
    survey_bytes = await survey_file.read() 

    # バイト列をopenpyxlで読み込む(load_workbookuだけでは、バイト列をそのまま渡すことになるためio.BytesIOを使用する)
    wb = load_workbook(io.BytesIO(survey_bytes), data_only=True)

    # 2. masterフォルダから固定ファイルを読み込む
    cell_mapping_file = MASTER_DIR / "forest_survey_cell_mapping.xlsx"
    check_rules_file = MASTER_DIR / "check_rules_forest_survey.csv"

    # 3. セル対応表を読み込む
    basic_information = load_basic_inf_cell_mapping(cell_mapping_file)
    tree_information = dict(load_tree_inf_cell_mapping(cell_mapping_file))
    start_row, end_row = load_tree_data_cell_mapping(cell_mapping_file)

    # 4. ルールを読み込む
    check_rules = load_check_rules(check_rules_file)
    basic_rules = get_basic_rules(check_rules)
    tree_rules = get_tree_rules(check_rules)

    # 5. シートごとにエラーチェックを実行する
    target_sheets = wb.sheetnames

    all_errors = []
    for sheet_name in target_sheets:
        ws = wb[sheet_name]

        basic_errors = check_basic_info(
            ws,
            sheet_name,
            basic_information,
            basic_rules,)

        tree_errors = check_tree_rows(
            ws,
            sheet_name,
            tree_information,
            start_row,
            end_row,
            tree_rules,)

        all_errors.extend(basic_errors)
        all_errors.extend(tree_errors)


    # 6. エラーがない場合のみGISサマリーを作成する
    gis_rows = []

    if not all_errors:
        for sheet_name in target_sheets:
            ws = wb[sheet_name]
            summary_row = create_gis_summary_row(
                ws,
                sheet_name,
                tree_information,
                start_row)
            gis_rows.append(summary_row)

    # 7. 結果を返す
    return {
        "has_error": len(all_errors) > 0,
        "error_count": len(all_errors),
        "error_log": all_errors,
        "gis_summary": gis_rows
    }